#authorised by Henry Tsai
import os
from logging import root
from pdb import line_prefix
from sysconfig import get_paths
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk, StringVar
import pandas as pd
from pandas import value_counts
import numpy as np
from ttkbootstrap import Style
import customtkinter  as ctk
from collections import defaultdict
import subprocess
import pymysql,openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Border,Side
from openpyxl.drawing import image
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.chart import LineChart,Reference
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph,CharacterProperties,ParagraphProperties,Font
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
#建立與mySQL連線資料
db_settings = { 
    "host": "127.0.0.1",
    "port": 3306,
    "user": "root",
    "password": "ROOT",
    "db": "nantou db",
    "charset": "utf8"
    }
try:
    conn = pymysql.connect(**db_settings)
except pymysql.err.OperationalError:
    db_settings.update({"host": "192.168.0.120","port": 3307})
    del db_settings["password"]
finally:
    conn = pymysql.connect(**db_settings)

#與mySQL建立連線，取出測試件項目工作表中的測試件名稱以及編號
with conn.cursor() as cursor:
    cursor.execute("SELECT `測試件項目`.測試件名稱, `測試件分項目`.測試項目_分項 FROM `測試件項目`, `測試件分項目` WHERE (`測試件項目`.編號 = `測試件分項目`.編號);")
result_subtestname = cursor.fetchall()
with conn.cursor() as cursor:
    cursor.execute("SELECT `測試件名稱`, `編號` FROM `測試件項目`;")
testname = cursor.fetchall()
with conn.cursor() as cursor:
    cursor.execute("SELECT `測試項目_分項`, `分項編號` FROM `測試件分項目`;")
testobj = cursor.fetchall()
result_subtestname=tuple((y, x) for x, y in result_subtestname)
testname_dict = dict((x,y)for x,y in testname)
# result_testobj = tuple((y,x)for x,y in testobj)
# print(testname_dict)
# print(result_testobj)
subtestname = defaultdict(list)
for i,j in result_subtestname:
    subtestname[j].append(i)
sorted(subtestname.items())
result=list(subtestname.keys())
# conn.close()

class output_mySQL(object):

    def __init__(self):  
        #建立資料輸入介面  
        self.root = ctk.CTk()
        # ctk.set_default_color_theme("green")  
        # 給主視窗設定標題內容  
        self.root.title("能力試驗結果匯出")  
        self.root.geometry('800x600')
        style = Style(theme='cyborg') 
        self.root.config(background='#323232') #設定背景色
        global testobj_values
        self.root.bind('<Return>', self.return_click)
        self.hellow_label = ctk.CTkLabel(
            self.root, 
            text = "能力試驗結果匯出",
            fg_color='#323232',
            text_font=('微軟正黑體',20),
            text_color="#00F5FF",
            width=260
            )
        self.label_year = ctk.CTkLabel(
            self.root, 
            text = "年份:", 
            fg_color='#323232',
            text_font=('微軟正黑體',18),
            text_color="#00F5FF",
            width=120
            )
        self.label_testname = ctk.CTkLabel(
            self.root, 
            text = "測試件名稱:", 
            fg_color='#323232',
            text_font=('微軟正黑體',18),
            text_color="#00F5FF",
            width=150
            )
        self.label_testnum = ctk.CTkLabel(
            self.root, 
            text = "年度第幾次:", 
            fg_color='#323232',
            text_font=('微軟正黑體',18),
            text_color="#00F5FF",
            width=150
            )
        # self.label_testobj = ctk.CTkLabel(
        #     self.root, 
        #     text = "能力試驗項目:", 
        #     fg_color='#323232',
        #     text_font=('微軟正黑體',18),
        #     text_color="#00F5FF",
        #     width=150
        #     )
        self.input_year = ctk.CTkEntry(
            self.root, 
            fg_color='#666666',
            text_color="#FFFFFF",
            width=150,height=35
            )
        self.input_testnum = ctk.CTkEntry(
            self.root, 
            fg_color='#666666',
            text_color="#FFFFFF",
            width=150,height=35
            )
        self.input_testname = ttk.Combobox(
            master = self.root, 
            values=result,
            width=15,height=40,
            state="readonly"
            )
        # self.input_testname.bind("<<ComboboxSelected>>", self.callback)
        # self.input_testobj = ttk.Combobox(
        #     master = self.root, 
        #     # textvariable=self.input_testname.get(),
        #     values=subtestname[self.input_testname.get()],
        #     width=15,height=40,
        #     state="readonly"
        #     )
        self.button_back=ctk.CTkButton(
            self.root, 
            command = self.back_interface, 
            text = "返回", 
            fg_color='#666666',
            width=180,height=40,
            text_font='微軟正黑體',
            text_color="#00F5FF"
            )
        self.button_back.pack()
        self.button_OK=ctk.CTkButton(
            self.root, 
            command = lambda: self.OK_interface(1), 
            text = "確定", 
            fg_color='#666666',
            width=180,height=40,
            text_font='微軟正黑體',
            text_color="#00F5FF"
            )
        self.button_back.pack()
        self.cc = ctk.CTkLabel(
            self.root, 
            fg_color="#323232",
            text='@Design by Henry Tsai',
            text_color="#8E8E8E",
            text_font="Calibri",
            width=170)

    
    def back_interface(self):
        self.root.destroy()
    
    def OK_interface(self,pending):
        input_year = self.input_year.get()  #輸入年份擷取
        input_testnum = self.input_testnum.get()    #今年第幾次擷取
        input_testname = testname_dict[self.input_testname.get()]
        input_testname = int(input_testname)
        input_testnum = int(input_testnum)
        # input_testobj = self.input_testobj.get()
        with conn.cursor() as cursor:
            srch_objnum = "SELECT `測試件數` FROM `測試件項目` WHERE `測試件項目`.`編號` = %d;"%(input_testname)
            cursor.execute(srch_objnum)
        objnum = cursor.fetchone()
        objnum = objnum[0]
        if input_year=="" or input_testnum=="":
            tk.messagebox.showinfo(title='南投署立醫院檢驗科', message='輸入不完全!請重新輸入!')
        else:
            with conn.cursor() as cursor:
                totalobj = """SELECT COUNT(*) 
                                FROM `測試件分項目`
                                WHERE `編號`= %d;"""%(input_testname)
                cursor.execute(totalobj)
            totalobj = cursor.fetchall()
            totalobj = int(totalobj[0][0])
            with conn.cursor() as cursor:
                if pending == 1:
                    srch_db="""SELECT `測試件結果`.`年份`,  `測試件結果`.`年度次數`, `測試件結果`.`測試件項目編號`, `測試件分項目`.`測試項目_分項`, `測試件結果`.`測試件序號`,`測試件結果`.`測試件結果`, `能力試驗結果`.`能力試驗數值` 
                            FROM `測試件結果`
                            JOIN `測試件項目`
                            ON `測試件結果`.`測試件項目編號` = `測試件項目`.`編號` 
                            JOIN `測試件分項目`
                            ON`測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`
                            JOIN `能力試驗結果`
                            ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                            WHERE `測試件結果`.`測試件分項目編號` IN  (SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d)
                            AND `測試件結果`.`年份` = %s 
                            AND `測試件結果`.`年度次數` = %d ;
                            """%(input_testname,input_year,input_testnum)
                elif pending != 1:
                    srch_db="""SELECT `測試件結果`.`年份`,  `測試件結果`.`年度次數`, `測試件結果`.`測試件項目編號`, `測試件分項目`.`測試項目_分項`, `測試件結果`.`測試件序號`,`測試件結果`.`測試件結果_備機`, `能力試驗結果`.`能力試驗數值` 
                            FROM `測試件結果`
                            JOIN `測試件項目`
                            ON `測試件結果`.`測試件項目編號` = `測試件項目`.`編號` 
                            JOIN `測試件分項目`
                            ON`測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`
                            JOIN `能力試驗結果`
                            ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                            WHERE `測試件結果`.`測試件分項目編號` IN  (SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d)
                            AND `測試件結果`.`年份` = %s 
                            AND `測試件結果`.`年度次數` = %d ;
                            """%(input_testname,input_year,input_testnum)
                cursor.execute(srch_db)
            name = cursor.fetchall()
            cnt = cursor.rowcount
            if cnt == 0:
                tk.messagebox.showinfo(title='南投署立醫院檢驗科', message='查無此筆資料!請確定年份或次數是否輸入正確!')
                self.clear_data()
                return
            # elif actualcnt != cnt:
            #     self.db_insufficient()
            
            name = pd.DataFrame(name)
            testobj_1 = name[3].unique()
            testobj_1.tolist()
            wb = Workbook()
            m,upsql = 0,0
            for q in testobj_1:
                with conn.cursor() as cursor:
                    if pending == 1:
                        srch_db="""SELECT `測試件結果`.`年份`,  `測試件結果`.`年度次數`, `測試件結果`.`測試件項目編號`, `測試件分項目`.`測試項目_分項`, `測試件結果`.`測試件序號`,`測試件結果`.`測試件結果`, `能力試驗結果`.`能力試驗數值`,`測試件結果`.`不等判讀` 
                            FROM `測試件結果`
                            JOIN `測試件項目`
                            ON `測試件結果`.`測試件項目編號` = `測試件項目`.`編號` 
                            JOIN `測試件分項目`
                            ON`測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`
                            JOIN `能力試驗結果`
                            ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                            WHERE `測試件結果`.`測試件分項目編號` IN  (SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s')
                            AND `測試件結果`.`年份` = %s 
                            AND `測試件結果`.`年度次數` = %d 
                            ORDER BY `測試件序號` ASC;
                            """%(input_testname,q,input_year,input_testnum)
                    elif pending != 1:
                        srch_db="""SELECT `測試件結果`.`年份`,  `測試件結果`.`年度次數`, `測試件結果`.`測試件項目編號`, `測試件分項目`.`測試項目_分項`, `測試件結果`.`測試件序號`,`測試件結果`.`測試件結果_備機`, `能力試驗結果`.`能力試驗數值` ,`測試件結果`.`不等判讀_備機`
                            FROM `測試件結果`
                            JOIN `測試件項目`
                            ON `測試件結果`.`測試件項目編號` = `測試件項目`.`編號` 
                            JOIN `測試件分項目`
                            ON`測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`
                            JOIN `能力試驗結果`
                            ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                            WHERE `測試件結果`.`測試件分項目編號` IN  (SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s')
                            AND `測試件結果`.`年份` = %s 
                            AND `測試件結果`.`年度次數` = %d 
                            ORDER BY `測試件序號` ASC;
                            """%(input_testname,q,input_year,input_testnum)
                    cursor.execute(srch_db)
                rawdata = cursor.fetchall()
                rawcnt = cursor.rowcount
                with conn.cursor() as cursor:
                    srch_clsi = """SELECT `能力試驗結果`.`測試件結果編號`, `測試件分項目`.`測試項目_分項`,`clsi規則`.`規則內容`, `clsi規則`.`實際數值`,`clsi規則`.`實際數值_1`
                                    FROM (`測試件結果` INNER JOIN `測試件分項目` ON `測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`)
                                    JOIN `能力試驗結果`
                                    ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                                    JOIN `clsi規則`
                                    ON `能力試驗結果`.`CLSI規則` = `clsi規則`.`編號`
                                    WHERE `能力試驗結果`.`測試件結果編號` IN  (
                                        SELECT `結果編號` FROM `測試件結果` WHERE`測試件結果`.`測試件分項目編號`IN(
                                            SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s')
                                        AND`測試件結果`.`年份`='%s'
                                        AND`測試件結果`.`年度次數`=%d);
                                """%(input_testname,q,input_year,input_testnum)
                    cursor.execute(srch_clsi)
                rules = cursor.fetchone()
                try:
                    aa = rules[2]
                except TypeError:
                    tk.messagebox.showerror(title='南投醫院檢驗科',message='尚未請組長上傳可容許範圍，請聯繫組長或資訊醫檢師!')
                    srch_presetclsi = """SELECT `clsi規則`.`規則內容`,`clsi規則`.`實際數值`,`clsi規則`.`實際數值_1`
                                    FROM `測試件分項目`
                                    JOIN  `clsi規則`
                                    ON `clsi規則`.`編號` = `測試件分項目`.`預設規則`
                                    WHERE `測試項目_分項` = '%s';"""%(q)
                with conn.cursor() as cursor:
                    cursor.execute(srch_presetclsi)
                prerules = cursor.fetchone()
                pre = prerules[0]
                if tk.messagebox.askyesno(title='南投醫院檢驗科',message='%s是否先使用預設可容許範圍:%s?'%(q,pre)):
                    aa = pre
                    rules = list(prerules)
                    rules.insert(0,"")
                    rules.insert(0,"")
                    print(rules)
                else:
                    return
                # print(rules[3])
                if m == 0:  #判斷是否新增一個新的excel檔
                    ws = wb.active
                    m += 1
                else:
                    ws = wb.create_sheet()
                ws.title = "%s_%d_%s_%s"%(input_year,input_testnum,self.input_testname.get(),q)
                # ws.title()
                title=["年份","年度次數","測試件項目","測試件分項目","測試件序號","測試件結果","目標值"]
                ws.append(title)
                #放入目標數值
                if rawcnt!= objnum:
                    rawdata = self.db_insufficient(result_db = rawdata, objnum = objnum)
                else: 
                    rawdata = pd.DataFrame(rawdata)
                print(rawdata)
                for j in range(0, objnum):
                    row_name = rawdata.iloc[j]
                    row_name = row_name.values.tolist()
                    if row_name[-1] == "=":
                        row_name.pop()
                        ws.append(row_name)
                    elif row_name[-1] == None:
                        ws.append(row_name)
                    else:
                        nonequal = row_name[-1]
                        row_name[5] = nonequal + str(row_name[5])
                        row_name.pop()
                        ws.append(row_name)
                #刪除不必要資訊
                ws.delete_cols(1,amount = 4)
                ws.delete_cols(4)
                ##計算與peer差異值
                ws['D1'] = "差異值"
                for x in range(2,objnum+2):
                    if ws['B' + str(x)].value == None:
                        tk.messagebox.showinfo(title='南投署立醫院檢驗科', message='測試件無結果!，無法自動計算可容許範圍')    
                    else:
                        actualmean = ws['B' + str(x)].value
                        peermean =  ws['C' + str(x)].value
                        try:
                            ws['D' + str(x)].value = actualmean - peermean
                        except TypeError:
                            continue
                ##新增標題
                ws.insert_rows(1)
                ws['A1'] = "%s年第%d次%s_%s"%(input_year,input_testnum,self.input_testname.get(),q)
                # ws.merge_cells('A1:B1')
                # ws.merge_cells('A23:B23')
                ws['E2'] = "可容許差異高值"
                ws['F2'] = "可容許差異低值"
                ws['G2'] = "差異百分比"
                ##計算可容許差異高/低值
                if "or" in aa:  #如果多於兩個變數
                    if "%" and "SD" in aa:  #如果同時出現SD跟%
                        percentage =[]
                        with conn.cursor() as cursor:   #取得SD值
                            srch_sd = """SELECT `測試件結果`.`測試件序號`,`能力試驗結果`.`測試件結果編號`, `能力試驗結果`.`能力試驗標準差`
                                        FROM (`測試件結果` INNER JOIN `測試件分項目` ON `測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`)
                                        JOIN `能力試驗結果`
                                        ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                                        WHERE `能力試驗結果`.`測試件結果編號` IN  (
                                            SELECT `結果編號` FROM `測試件結果` WHERE`測試件結果`.`測試件分項目編號`IN(
                                                SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s')
                                                AND `測試件結果`.`年份` = '%s' 
                                                AND `測試件結果`.`年度次數` = %d);
                            """%(input_testname,q,input_year,input_testnum)
                            cursor.execute(srch_sd)
                        sd = cursor.fetchall()
                        sd_row = cursor.rowcount
                        if sd_row != objnum:
                            sd = self.sd_insufficient(testnamenum=input_testname,testobj=q,year=input_year,testnum=input_testnum,objnum=objnum)
                        else:
                            sd = pd.DataFrame(sd)
                        sd = sd.values.tolist()
                        for i in range(3, objnum + 3):
                            if ws['B' + str(i)].value == None:
                                percentage.append(None)
                                continue
                            else:
                                realamount_1 = rules[3]
                                realamount_2 = rules[4]
                                realamount_1 = ws['C' + str(i)].value * realamount_1
                                realamount_2 = sd[i-3][2] * realamount_1
                                if realamount_1 > realamount_2:
                                    ws['E' + str(i)].value = realamount_1
                                    ws['F' + str(i)].value = (-realamount_1)
                                    try:
                                        ws['G' + str(i)].value = ws['D' + str(i)].value / ws['E' + str(i)].value
                                    except TypeError:
                                        ws['G' + str(i)].value = "N/A"
                                    ws['G' + str(i)].number_format = "0.00%"
                                else: 
                                    ws['E' + str(i)].value = realamount_2
                                    ws['F' + str(i)].value = (-realamount_2)
                                    try:
                                        ws['G' + str(i)].value = ws['D' + str(i)].value / ws['E' + str(i)].value
                                    except TypeError:
                                        ws['G' + str(i)].value = "N/A"
                                    ws['G' + str(i)].number_format = "0.00%"
                                if ws['G' + str(i)].value != "N/A":
                                    percent = ws['G' + str(i)].value * 100
                                else:
                                    percent = "N/A"
                                percentage.append(percent)
                    elif "%" in aa:
                        percentage =[]
                        for z in range(3, objnum + 3):
                            if ws['B' + str(z)].value == None:
                                percentage.append(None)
                                continue
                            else:
                                realamount_1 = rules[3]
                                realamount_2 = rules[4]
                                realamount_1 = ws['C' + str(z)].value * realamount_1
                                if realamount_1 > realamount_2:
                                    ws['E' + str(z)].value = realamount_1
                                    ws['F' + str(z)].value = (-realamount_1)
                                    try:
                                        ws['G' + str(z)].value = ws['D' + str(z)].value / ws['E' + str(z)].value
                                    except TypeError:
                                        ws['G' + str(z)].value = "N/A"
                                    ws['G' + str(z)].number_format = "0.00%"
                                else: 
                                    ws['E' + str(z)].value = realamount_2
                                    ws['F' + str(z)].value = (-realamount_2)
                                    try:
                                        ws['G' + str(z)].value = ws['D' + str(z)].value / ws['E' + str(z)].value
                                    except TypeError:
                                        ws['G' + str(z)].value = "N/A"
                                    ws['G' + str(z)].number_format = "0.00%"
                                if ws['G' + str(z)].value != "N/A":
                                    percent = ws['G' + str(z)].value * 100
                                else:
                                    percent = "N/A"
                                percentage.append(percent)
                    elif "SD" in aa:
                        percentage =[]
                        with conn.cursor() as cursor:   #取得SD值
                            srch_sd = """SELECT `測試件結果`.`測試件序號`,`能力試驗結果`.`測試件結果編號`, `能力試驗結果`.`能力試驗標準差`
                                        FROM (`測試件結果` INNER JOIN `測試件分項目` ON `測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`)
                                        JOIN `能力試驗結果`
                                        ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                                        WHERE `能力試驗結果`.`測試件結果編號` IN  (
                                            SELECT `結果編號` FROM `測試件結果` WHERE`測試件結果`.`測試件分項目編號`IN(
                                                SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s')
                                                AND `測試件結果`.`年份` = '%s' 
                                                AND `測試件結果`.`年度次數` = %d);
                            """%(input_testname,q,input_year,input_testnum)
                            cursor.execute(srch_sd)
                        sd = cursor.fetchall()
                        sd_row = cursor.rowcount
                        if sd_row != objnum:
                            sd = self.sd_insufficient(testnamenum=input_testname,testobj=q,year=input_year,testnum=input_testnum,objnum=objnum)
                        else:
                            sd = pd.DataFrame(sd)
                        sd = sd.values.tolist()
                        for i in range(3, objnum + 3):
                            if ws['B' + str(i)].value == None:
                                percentage.append(None)
                                continue
                            else:
                                realamount_1 = rules[3]
                                realamount_2 = rules[4]
                                realamount_1 = sd[i-3][2] * realamount_1
                                if realamount_1 > realamount_2:
                                    ws['E' + str(i)].value = realamount_1
                                    ws['F' + str(i)].value = (-realamount_1)
                                    try:
                                        ws['G' + str(i)].value = ws['D' + str(i)].value / ws['E' + str(i)].value
                                    except TypeError:
                                        ws['G' + str(i)].value = "N/A"
                                    ws['G' + str(i)].number_format = "0.00%"
                                else: 
                                    ws['E' + str(i)].value = realamount_2
                                    ws['F' + str(i)].value = (-realamount_2)
                                    try:
                                        ws['G' + str(i)].value = ws['D' + str(i)].value / ws['E' + str(i)].value
                                    except TypeError:
                                        ws['G' + str(i)].value = "N/A"
                                    ws['G' + str(i)].number_format = "0.00%"
                                if ws['G' + str(i)].value != "N/A":
                                    percent = ws['G' + str(i)].value * 100
                                else:
                                    percent = "N/A"
                                percentage.append(percent)
                elif "SD" in aa:   #利用標準差計算高低值
                    percentage =[]
                    realamount = rules[3]
                    with conn.cursor() as cursor:
                        srch_sd = """SELECT `測試件結果`.`測試件序號`,`能力試驗結果`.`測試件結果編號`, `能力試驗結果`.`能力試驗標準差`
                                    FROM (`測試件結果` INNER JOIN `測試件分項目` ON `測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`)
                                    JOIN `能力試驗結果`
                                    ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                                    WHERE `能力試驗結果`.`測試件結果編號` IN  (
                                        SELECT `結果編號` FROM `測試件結果` WHERE`測試件結果`.`測試件分項目編號`IN(
                                            SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s')
                                            AND `測試件結果`.`年份` = '%s' 
                                            AND `測試件結果`.`年度次數` = %d);
                        """%(input_testname,q,input_year,input_testnum)
                        cursor.execute(srch_sd)
                    sd = cursor.fetchall()
                    sd_row = cursor.rowcount
                    if sd_row != objnum:
                        sd = self.sd_insufficient(testnamenum=input_testname,testobj=q,year=input_year,testnum=input_testnum,objnum=objnum)
                    else:
                        sd = pd.DataFrame(sd)
                    sd = sd.values.tolist()
                    for i in range(3,objnum + 3):
                        if ws['B' + str(i)].value == None:
                            percentage.append(None)
                            continue
                        else:
                            ws['E' + str(i)].value = sd[i-3][2] * realamount
                            ws['F' + str(i)].value = sd[i-3][2] * (-realamount)
                            if ws['E' + str(i)].value != 0:
                                try:
                                    ws['G' + str(i)].value = ws['D' + str(i)].value / ws['E' + str(i)].value
                                except TypeError:
                                    ws['G' + str(i)].value = "N/A"
                                ws['G' + str(i)].number_format = "0.00%"
                            else:
                                ws['G' + str(i)].value = 0
                            if ws['G' + str(i)].value !="N/A":
                                percent = ws['G' + str(i)].value * 100
                            else:
                                percent = "N/A"
                            percentage.append(percent)
                elif "%" in aa: #利用百分比計算高低值
                    percentage =[]
                    realamount = rules[3]
                    for i in range(3,objnum + 3):
                        if ws['B' + str(i)].value == None:
                            percentage.append(None)
                            continue
                        else:
                            ws['E' + str(i)].value = ws['C' + str(i)].value * realamount
                            ws['F' + str(i)].value = ws['C' + str(i)].value * (-realamount)
                            try:
                                ws['G' + str(i)].value = ws['D' + str(i)].value / ws['E' + str(i)].value
                            except TypeError:
                                ws['G' + str(i)].value = "N/A"
                                percent = "N/A"
                            else:
                                ws['G' + str(i)].number_format = "0.00%"
                                percent = ws['G' + str(i)].value * 100
                            percentage.append(percent)
                else:   #高低值為固定數值
                    percentage =[]
                    realamount = rules[3]
                    for i in range(3,objnum + 3):
                        if ws['B' + str(i)].value == None:
                            percentage.append(None)
                            continue
                        else:
                            ws['E' + str(i)].value = realamount
                            ws['F' + str(i)].value = (-realamount)
                            try:
                                ws['G' + str(i)].value = ws['D' + str(i)].value / ws['E' + str(i)].value
                            except TypeError:
                                ws['G' + str(i)].value = "N/A"
                            else:
                                ws['G' + str(i)].number_format = "0.00%"
                                percent = ws['G' + str(i)].value * 100
                            percentage.append(percent)
                
                ft1 = openpyxl.styles.Font(name='標楷體', size=20)  #標題字體設定
                ft2 = openpyxl.styles.Font(name='標楷體', size=12)  #版次字體設定
                ft3 = openpyxl.styles.Font(name='標楷體', size=10)  #表單編號字體設定
                ft4 = openpyxl.styles.Font(name='Times New Roman', size=11)  #內文字體設定
                ##標題置中對齊 & 自動適配欄寬
                for col in range(1,8):
                    char = get_column_letter(col)
                    ws[char + str(2)].font = ft3
                    for row in range(2,objnum + 3):
                        ws[char + str(row)].alignment = Alignment(horizontal='center',vertical='center')
                        if row == 2:
                            pass
                        else:
                            ws[char + str(row)].font = ft4
                    ws.column_dimensions[get_column_letter(col)].auto_size = True
                ##新增表頭"能力試驗可容許結果"
                ws.insert_rows(1,2) #新增兩行
                ws.merge_cells("A1:A2") #圖片欄位合併
                self.insert_img(worksheet = ws,img='logo1.png')    #新增置中圖片
                ws["B1"].value = "能力試驗可容許結果"   #新增標題
                ws["B1"].font = ft1
                ws["F1"].value = "版次:22.1版"  #新增版次
                ws["F1"].font = ft2
                ws["F2"].value = "表單編號:R-QP05063-010"  #新增表單編號
                ws["F2"].font = ft3
                ws.merge_cells("B1:E2") #標題欄位合併
                ws['B1'].alignment = Alignment(horizontal='center',vertical='center')   #標題置中
                ws.merge_cells("F1:G1") #版次欄位合併
                ws['F1'].alignment = Alignment(horizontal='center',vertical='center')   #版次置中
                ws.merge_cells("F2:G2") #表單編號欄位合併
                ws['F2'].alignment = Alignment(horizontal='center',vertical='center')   #表單編號置中
                ws.row_dimensions[1].height = 24    #設定行高
                ws.row_dimensions[2].height = 24    #設定行高
                ws.move_range('A3',cols=1)  #將能力試驗名稱移至A3
                ws['A3'].value = "能力試驗:"
                ws['A3'].alignment = Alignment(horizontal='right',vertical='center')    #能力試驗靠右對齊
                ws['A3'].font = ft2
                ws.merge_cells("B3:D3")
                ws['E3'].value = "可容許範圍:"
                ws['E3'].font = ft2
                ws['F3'].value = aa
                ws['F3'].font = ft4
                ##繪製差異圖表
                chart = LineChart()
                chart.title = ws['B3'].value    #圖表標題
                cp = CharacterProperties(ea= Font(typeface='標楷體'), sz = 1400, b = False) #設定標題字型
                lp = CharacterProperties(ea= Font(typeface='標楷體'), sz = 1000, b = False) #設定圖例字型
                pp = ParagraphProperties(defRPr=lp)
                rlp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
                chart.title.tx.rich.p[0].pPr.defRPr = cp     #標題設定
                chart.legend.textProperties = rlp       #圖例設定
                # chart.legend.textProperties.rich.p[0].pPr.defRPr = lp    #圖例設定
                chart.legend.position = "b"     #設定圖例放置位置
                ydata = Reference(ws, min_col=4, min_row=4, max_col=6, max_row=objnum+4)
                xvalue = Reference(ws, min_col=1, min_row=5, max_col=1, max_row=objnum+5)
                chart.add_data(ydata, titles_from_data=True)
                chart.set_categories(xvalue)
                s1 = chart.series[1]
                s1.marker.symbol = "circle"
                s1.marker.graphicalProperties.solidFill = "D44E2A" # Marker filling
                s1.marker.graphicalProperties.line.solidFill = "D44E2A" # Marker outline
                s1.graphicalProperties.line.solidFill = "D44E2A"
                s2 = chart.series[0]
                s2.marker.symbol = "circle"
                s2.marker.graphicalProperties.solidFill = "3580EF" # Marker filling
                s2.marker.graphicalProperties.line.solidFill = "3580EF" # Marker outline
                s2.graphicalProperties.line.solidFill = "3580EF"
                s3 = chart.series[2]
                s3.marker.symbol = "circle"
                s3.marker.graphicalProperties.solidFill = "71B62C" # Marker filling
                s3.marker.graphicalProperties.line.solidFill = "71B62C" # Marker outline
                s3.graphicalProperties.line.solidFill = "71B62C"
                ##設定絕對座標
                p2e = pixels_to_EMU
                position = XDRPoint2D(p2e(30), p2e(200))
                size = XDRPositiveSize2D(p2e(560), p2e(260))
                chart.anchor = AbsoluteAnchor(pos=position, ext=size)
                ws.add_chart(chart)
                ##計算差異百分比後自動回傳mySQL
                #先檢查是否為自己判定數值
                if ws['E3'].value =="計算複雜，請自行計算":
                    ws['C1'].value = aa
                    tk.messagebox.showinfo(title='南投署立醫院檢驗科', message='%s可容許範圍判定大於兩個變數，請自行判斷'%(q))
                else:
                    if pending == 1:
                        if upsql==0:
                            if tk.messagebox.askyesno(title='南投署立醫院檢驗科', message="""是否上傳計算後可容許百分比?
                            項目: %s"""%(q)):
                                upsql = 1
                                with conn.cursor() as cursor:
                                    srch_num ="""SELECT `測試件結果`.`結果編號` 
                                                FROM `測試件結果`
                                                WHERE `測試件結果`.`測試件分項目編號` IN  (
                                                    SELECT `分項編號` 
                                                    FROM `測試件分項目` 
                                                    WHERE `編號` = %d 
                                                    AND `測試項目_分項` = '%s')
                                                AND `測試件結果`.`年份` = %s 
                                                AND `測試件結果`.`年度次數` = %d ;"""%(input_testname,q,input_year,input_testnum)
                                    cursor.execute(srch_num)
                                rowcnt = cursor.rowcount
                                if rowcnt != objnum:
                                    srch_number = self.update_insufficient(testname=input_testname,testobj=q,year=input_year,testnum=input_testnum,objnum=objnum)
                                else:
                                    srch_number = [srch_num_1[0] for srch_num_1 in cursor.fetchall()]
                                for l in range(0,len(srch_number)):
                                    if percentage[l] == None or percentage[l]=="N/A":
                                        continue
                                    else:
                                        with conn.cursor() as cursor:
                                            input_percent="""UPDATE `能力試驗結果` SET`差異百分比`=%.5f WHERE `測試件結果編號`='%s';"""%(percentage[l],srch_number[l])
                                            # print(input_percent)
                                            cursor.execute(input_percent)
                                        conn.commit()
                            else:
                                upsql = 2
                                pass
                        elif upsql == 1:
                            with conn.cursor() as cursor:
                                srch_num ="""SELECT `測試件結果`.`結果編號` 
                                            FROM `測試件結果`
                                            WHERE `測試件結果`.`測試件分項目編號` IN  (
                                                SELECT `分項編號` 
                                                FROM `測試件分項目` 
                                                WHERE `編號` = %d 
                                                AND `測試項目_分項` = '%s')
                                            AND `測試件結果`.`年份` = %s 
                                            AND `測試件結果`.`年度次數` = %d ;"""%(input_testname,q,input_year,input_testnum)
                                cursor.execute(srch_num)
                            rowcnt = cursor.rowcount
                            if rowcnt != objnum:
                                srch_number = self.update_insufficient(testname=input_testname,testobj=q,year=input_year,testnum=input_testnum,objnum=objnum)
                            else:
                                srch_number = [srch_num_1[0] for srch_num_1 in cursor.fetchall()]
                            for l in range(0,len(srch_number)):
                                if percentage[l] == None or percentage[l]=="N/A":
                                    continue
                                else:
                                    with conn.cursor() as cursor:
                                        input_percent="""UPDATE `能力試驗結果` SET`差異百分比`=%.5f WHERE `測試件結果編號`='%s';"""%(percentage[l],srch_number[l])
                                        # print(input_percent)
                                        cursor.execute(input_percent)
                                    conn.commit()
                        elif upsql == 2:
                            pass
                    elif pending != 1:
                        pass
                        # tk.messagebox.showinfo(title='南投署立醫院檢驗科', message='備機不須上傳可容許範圍!')
                ###擷取近五次資料
                ##所有能力試驗結果，每年測試次數皆為2或3
                with conn.cursor() as cursor:   #取得近五年
                    srch_year="""SELECT DISTINCT `測試件結果`.`年份`,`測試件結果`.`年度次數`
                    FROM (`測試件結果` INNER JOIN `測試件分項目` ON `測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`)
                    JOIN `能力試驗結果`
                    ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                    WHERE `能力試驗結果`.`測試件結果編號` IN  (
                        SELECT `結果編號` FROM `測試件結果` WHERE`測試件結果`.`測試件分項目編號`IN(
                            SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s'))
                    ORDER BY `測試件結果`.`年份` DESC, `測試件結果`.`年度次數` DESC
                    LIMIT %d;"""%(input_testname,q,objnum)
                    cursor.execute(srch_year)
                year = cursor.fetchall()
                title =['測試件名稱']
                for i in range(len(year)-1,-1,-1):
                    title.append('%s年第%d次'%(year[i][0],year[i][1]))
                # print(title)
                with conn.cursor() as cursor:
                    get_past = """SELECT `能力試驗結果`.`測試件結果編號`, `測試件結果`.`年份`,`測試件結果`.`年度次數`,`測試件結果`.`測試件序號`,`能力試驗結果`.`差異百分比`
                                FROM (`測試件結果` INNER JOIN `測試件分項目` ON `測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`)
                                JOIN `能力試驗結果`
                                ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                                WHERE `能力試驗結果`.`測試件結果編號` IN  (
                                    SELECT `結果編號` FROM `測試件結果` WHERE`測試件結果`.`測試件分項目編號`IN(
                                        SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s'))
                                ORDER BY `測試件結果`.`年份` DESC, `測試件結果`.`年度次數` DESC, `測試件結果`.`測試件序號` ASC
                                LIMIT %d;"""%(input_testname,q,objnum*5)
                    cursor.execute(get_past)
                ans = cursor.fetchall()
                anscount = cursor.rowcount
                if anscount % objnum !=0:
                    msg = "%s年%s項目不足!"%(int(year[i][0]),q)
                    tk.messagebox.showerror(title='南投署立醫院檢驗科', message = msg)
                    ans = self.backup_interface(haveyear=year,objnum=objnum,testobj=q)
                    ans.reset_index()
                else:
                    ans = pd.DataFrame(ans)
                    ans.reset_index()
                ws['A25'] = "能力試驗歷次監控"
                ws.merge_cells('A25:B25')
                ws.append(title)
                for i in range(5,5+objnum):    #新增能力試驗序號
                    ws['A' + str(i+22)].value = ws['A'+str(i)].value  #複製測試件項目
                    ws['A' + str(i+22)].font = ft4
                ws['A'+ str(27 + objnum)] = "平均"
                if len(title) == 6:
                    n = 0   #設定pd變數
                    for col in range(6,1,-1):  #五年固定
                        for row in range(27,objnum+27): #利用測試件數建立迴圈
                            char = get_column_letter(col)
                            ws[char + str(row)].value = ans.iat[n, 4]
                            ws[char + str(row)].number_format = '0.00'
                            ws[char + str(row)].alignment = Alignment(horizontal='center',vertical='center')
                            ws[char + str(row)].font = ft4
                            n+=1
                        # ws[get_column_letter(col) + str(objnum+25)].value = "=AVERAGE(B25:B29)"
                        ws[get_column_letter(col) + str(objnum+27)].value = "=ROUND(AVERAGE(%s%d:%s%d),2)"%(get_column_letter(col),27,get_column_letter(col),objnum+26)
                        ws[get_column_letter(col) + str(objnum+27)].alignment = Alignment(horizontal='center',vertical='center')
                        ws[get_column_letter(col) + str(objnum+27)].font = ft4
                        # ws[get_column_letter(col) + str(objnum+25)].number_format = "0.00%"
                else:
                    # tk.messagebox.showinfo(title='南投署立醫院檢驗科', message='歷年能力試驗不足五年!')
                    n = 0   #設定pd變數
                    for col in range(len(title),1,-1):  #不確定幾年
                        for row in range(27,objnum+27): #利用測試件數建立迴圈
                            char = get_column_letter(col)
                            ws[char + str(row)].value = ans.iat[n, 4]
                            ws[char + str(row)].number_format = '0.00'
                            ws[char + str(row)].alignment = Alignment(horizontal='center',vertical='center')
                            ws[char + str(row)].font = ft4
                            n+=1
                        # ws[clsget_column_letter(col) + str(objnum+25)].value = "=AVERAGE(B25:B29)"
                        ws[get_column_letter(col) + str(objnum+27)].value = "=ROUND(AVERAGE(%s%d:%s%d),2)"%(get_column_letter(col),27,get_column_letter(col),objnum+26)
                        ws[get_column_letter(col) + str(objnum+27)].alignment = Alignment(horizontal='center',vertical='center')
                        ws[get_column_letter(col) + str(objnum+27)].font = ft4
                        # ws[get_column_letter(col) + str(objnum+25)].number_format = "0.00%"
                ##歷次監控文字格式設定
                ws['A25'].font = ft2
                pastfont = "A26:%s26"%(get_column_letter(objnum))
                for row in ws[pastfont]:
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center',vertical='center')
                        cell.font = ft3
                ws['A'+ str(27 + objnum)].alignment = Alignment(horizontal='center',vertical='center')
                ws['A'+ str(27 + objnum)].font = ft3
                ##框線設定
                left, right, top, bottom = [Side(style='thin',color='000000')]*4  #新增框線
                border = Border(left=left, right=right, top=top, bottom=bottom)
                for row in ws["A1:G9"]:
                    for cell in row:
                        cell.border = border
                borderng = "A25:F%d"%(objnum+27)
                for row in ws[borderng]:
                    for cell in row:
                        cell.border = border
                ##繪製近五年圖表
                chart1 = LineChart()
                chart1.title = ws['B3'].value    #圖表標題
                cp = CharacterProperties(ea= Font(typeface='標楷體'), sz = 1400, b = False) #設定標題字型
                lp = CharacterProperties(ea= Font(typeface='標楷體'), sz = 1000, b = False) #設定圖例字型
                pp = ParagraphProperties(defRPr=lp)
                rlp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
                chart1.title.tx.rich.p[0].pPr.defRPr = cp     #標題設定
                chart1.legend.textProperties = rlp
                chart1.legend.position = "b"     #設定圖例放置位置
                chart1.y_axis.scaling.min = -100      # 設置y軸座標最小的值
                chart1.y_axis.majorUnit = 20       
                chart1.y_axis.scaling.max = 100    
                chart1.y_axis.numFmt = "General"

                ydata = Reference(ws, min_col=1, min_row=27, max_col=6, max_row=objnum+27)
                xvalue = Reference(ws, min_col=2, max_col=6, min_row=26, max_row=26)
                chart1.add_data(ydata, from_rows = True, titles_from_data=True)
                chart1.set_categories(xvalue)
                s1 = chart1.series[1]
                s1.marker.symbol = "circle"
                s1.marker.size = 7
                s1.marker.graphicalProperties.solidFill = "D44E2A" # Marker filling
                s1.marker.graphicalProperties.line.solidFill = "D44E2A"
                s1.graphicalProperties.line.noFill = True 
                s2 = chart1.series[0]
                s2.marker.symbol = "triangle"
                s2.marker.size = 7
                s2.marker.graphicalProperties.solidFill = "3580EF" # Marker filling
                s2.marker.graphicalProperties.line.solidFill = "3580EF"
                s2.graphicalProperties.line.noFill = True 
                s3 = chart1.series[2]
                s3.marker.symbol = "square"
                s3.marker.size = 7  
                s3.marker.graphicalProperties.solidFill = "6FB7B7" # Marker filling
                s3.marker.graphicalProperties.line.solidFill = "6FB7B7"
                s3.graphicalProperties.line.noFill = True 
                if objnum == 2:
                    s3.graphicalProperties.line.noFill = False
                    s3.graphicalProperties.line.solidfill = "6FB7B7"
                elif  objnum == 3:
                    s4 = chart1.series[3]
                    s4.marker.symbol = "diamond"
                    s4.marker.size = 7
                    s4.graphicalProperties.line.solidfill = "71B62C"
                elif objnum == 5 :
                    s4 = chart1.series[3]
                    s4.marker.symbol = "diamond"
                    s4.marker.size = 7
                    s4.marker.graphicalProperties.solidFill = "71B62C" # Marker filling
                    s4.marker.graphicalProperties.line.solidFill = "71B62C"
                    s4.graphicalProperties.line.noFill = True 
                    s5 = chart1.series[4]
                    s5.marker.symbol = "triangle"
                    s5.marker.size = 7
                    s5.marker.graphicalProperties.solidFill = "004B97" # Marker filling
                    s5.marker.graphicalProperties.line.solidFill = "004B97"
                    s5.graphicalProperties.line.noFill = True 
                    s6 = chart1.series[5]
                    s6.marker.symbol = "triangle"
                    s6.marker.size = 7
                    s6.graphicalProperties.solidFill = "FF9224" # Marker filling
                    s6.graphicalProperties.line.solidFill = "FF9224"
                ##設定圖表絕對位置
                p2e = pixels_to_EMU
                position = XDRPoint2D(p2e(30), p2e(650))
                size = XDRPositiveSize2D(p2e(560), p2e(260))
                chart1.anchor = AbsoluteAnchor(pos=position, ext=size)
                ws.add_chart(chart1)
                
                if pending == 1:
                    hhh=""
                elif pending!= 1:
                    hhh="(備機)"
                wb.save("%s年第%d次%s%s.xlsx"%(input_year,input_testnum,self.input_testname.get(),hhh))   #xlsx檔案存檔
                # wb.save("//10.0.35.9/95_report_RAW/能力試驗可容許範圍分析/程式自動暫存/%s年第%d次%s%s.xlsx"%(input_year,input_testnum,self.input_testname.get(),self.input_testobj.get(),hhh))   #xlsx檔案存檔
                filepath=".//%s年第%d次%s%s.xlsx"%(input_year,input_testnum,self.input_testname.get(),hhh)
        if os.path.isfile(filepath):
            tk.messagebox.showinfo(title='南投署立醫院檢驗科', message='檔案新增成功!')
            if pending == 1:
                if tk.messagebox.askyesno(title='南投署立醫院檢驗科', message='是否新增備機資料?'):
                    self.OK_interface(2)
                if tk.messagebox.askyesno(title='南投署立醫院檢驗科', message='是否開啟檔案?'):
                    command = "start " + filepath
                    subprocess.run(command, shell=True)
            else:
                self.clear_data()
                return
                # conn.close()
        else:
            tk.messagebox.showinfo(title='南投署立醫院檢驗科', message='檔案新增失敗QQ')
            # conn.close()
    
    def db_insufficient(self,result_db,objnum):
        db = pd.DataFrame(result_db)
        db = db.set_axis(["年份","年度次數","測試件項目編號","測試項目_分項","測試件序號","測試件結果","能力試驗數值","不等判讀"],axis=1,inplace=False)
        db.set_index("測試件序號")
        new_index = pd.Index(np.arange(1,objnum+1,1), name="測試件序號")
        db = db.set_index("測試件序號").reindex(new_index)
        db.reset_index(inplace = True)
        db = db.replace({np.nan:None})
        db = db[["年份","年度次數","測試件項目編號","測試項目_分項","測試件序號","測試件結果","能力試驗數值","不等判讀"]]
        return db
    
    def update_insufficient(self,testname,testobj,year,testnum,objnum):
        with conn.cursor() as cursor:
            srch_num ="""SELECT `測試件結果`.`測試件序號`,`測試件結果`.`結果編號`
                        FROM `測試件結果`
                        WHERE `測試件結果`.`測試件分項目編號` IN  (
                            SELECT `分項編號` 
                            FROM `測試件分項目` 
                            WHERE `編號` = %d 
                            AND `測試項目_分項` = '%s')
                        AND `測試件結果`.`年份` = %s 
                        AND `測試件結果`.`年度次數` = %d ;"""%(testname,testobj,year,testnum)
            cursor.execute(srch_num)
        up = cursor.fetchall()
        up = pd.DataFrame(up)
        up = up.set_axis(["測試件序號","結果編號"],axis=1,inplace=False)
        up.set_index("測試件序號")
        new_index = pd.Index(np.arange(1,objnum+1,1), name="測試件序號")
        up = up.set_index("測試件序號").reindex(new_index)
        up.reset_index(inplace = True)
        up = up.replace({np.nan:None})
        upans = up["結果編號"].tolist()
        return upans 

    def sd_insufficient(self,testnamenum,testobj,year,testnum,objnum): #解決如果能力試驗不足五項，但需要利用SD計算可容許範圍時
        with conn.cursor()as cursor:
            srch_sd = """SELECT `測試件結果`.`測試件序號`,`能力試驗結果`.`測試件結果編號`, `能力試驗結果`.`能力試驗標準差`
                        FROM (`測試件結果` INNER JOIN `測試件分項目` ON `測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`)
                        JOIN `能力試驗結果`
                        ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                        WHERE `能力試驗結果`.`測試件結果編號` IN  (
                            SELECT `結果編號` FROM `測試件結果` WHERE`測試件結果`.`測試件分項目編號`IN(
                                SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s')
                                AND `測試件結果`.`年份` = '%s' 
                                AND `測試件結果`.`年度次數` = %d);
                        """%(testnamenum,testobj,year,testnum)
            cursor.execute(srch_sd)
        ans = cursor.fetchall()
        ans = pd.DataFrame(ans)
        ans = ans.set_axis(["測試件序號","測試件結果編號","能力試驗標準差"],axis=1,inplace=False)
        ans.set_index("測試件序號")
        new_index = pd.Index(np.arange(1,objnum+1,1), name="測試件序號")
        ans = ans.set_index("測試件序號").reindex(new_index)
        ans.reset_index(inplace = True)
        ans = ans.replace({np.nan:None})
        return ans
    
    def backup_interface(self,haveyear,objnum,testobj): #歷年不足五次救援
        input_testname = testname_dict[self.input_testname.get()]
        input_testname = int(input_testname)
        # print(haveyear)
        finalans = pd.DataFrame()
        for i in range(0,len(haveyear)):
            with conn.cursor() as cursor:
                get_past = """SELECT `能力試驗結果`.`測試件結果編號`, `測試件結果`.`年份`,`測試件結果`.`年度次數`,`測試件結果`.`測試件序號`,`能力試驗結果`.`差異百分比`
                                FROM (`測試件結果` INNER JOIN `測試件分項目` ON `測試件結果`.`測試件分項目編號` = `測試件分項目`.`分項編號`)
                                JOIN `能力試驗結果`
                                ON `測試件結果`.`結果編號` = `能力試驗結果`.`測試件結果編號`
                                WHERE `能力試驗結果`.`測試件結果編號` IN  (
                                    SELECT `結果編號` FROM `測試件結果` WHERE`測試件結果`.`測試件分項目編號`IN(
                                        SELECT `分項編號` FROM `測試件分項目` WHERE `編號` = %d AND `測試項目_分項` = '%s'))
                                AND `測試件結果`.`年份`= %s
                                AND `測試件結果`.`年度次數`=%d
                                ORDER BY `測試件結果`.`年份` DESC, `測試件結果`.`年度次數` DESC, `測試件結果`.`測試件序號` ASC
                                LIMIT %d;"""%(input_testname,testobj,int(haveyear[i][0]),int(haveyear[i][1]),objnum)
                cursor.execute(get_past)
            ans = cursor.fetchall()
            ansrowcount = cursor.rowcount
            ans = pd.DataFrame(ans)
            if ansrowcount != objnum:
                ans = ans.set_axis(["測試件結果編號","年份","年度次數","測試件序號","差異百分比"],axis=1,inplace=False)
                ans.set_index("測試件序號")
                new_index = pd.Index(np.arange(1,objnum+1,1), name="測試件序號")
                ans = ans.set_index("測試件序號").reindex(new_index)
                ans = ans.fillna({"年份":int(haveyear[i][0]),"年度次數":int(haveyear[i][1])})
                ans.reset_index(inplace = True)
                ans = ans.replace({np.nan:None})
                finalans = pd.concat([finalans,ans],axis=0)
            else:
                ans = pd.DataFrame(ans)
                ans = ans.set_axis(["測試件結果編號","年份","年度次數","測試件序號","差異百分比"],axis=1,inplace=False)
                finalans = pd.concat([finalans,ans],axis=0)
        return finalans
    
    def insert_img(self, worksheet, img):
        img = image.Image(img)
        p2e = pixels_to_EMU
        position = XDRPoint2D(p2e(15), p2e(3))
        size = XDRPositiveSize2D(p2e(60), p2e(60))
        img.anchor = AbsoluteAnchor(pos=position, ext=size)
        worksheet.add_image(img)

    def clear_data(self):
        self.input_year.delete(0,"end")
        self.input_testname.set("")
        self.input_testnum.delete(0,"end")
    def gui_arrang(self):
        self.hellow_label.place(relx=0, rely=0.1, anchor=tk.W)
        self.label_year.place(relx=0.31, rely=0.2, anchor=tk.W)
        self.label_testnum.place(relx=0.25,rely=0.3,anchor=tk.W)
        self.label_testname.place(relx=0.25,rely=0.4,anchor=tk.W)
        self.input_year.place(relx=0.43, rely=0.2, anchor=tk.W)
        self.input_testnum.place(relx=0.43, rely=0.3, anchor=tk.W)
        self.input_testname.place(relx=0.45,rely=0.4,anchor=tk.W)
        self.button_back.place(relx=0.35,rely=0.8,anchor=tk.CENTER)
        self.button_OK.place(relx=0.65,rely=0.8,anchor=tk.CENTER)
        self.cc.place(relx=1, rely=1,anchor=tk.SE)
    def return_click(self, event):  #按Enter鍵自動連結登入
        self.OK_interface(1)
def main():  
    L = output_mySQL()
    L.gui_arrang()
    # 主程式執行  
    tk.mainloop()

if __name__ == '__main__':  
    main()  